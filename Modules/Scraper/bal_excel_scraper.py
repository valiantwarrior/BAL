from openpyxl import load_workbook

# E5
ROUTINE_NAME_ROW = 5
ROUTINE_NAME_COL = 5

# S5
AUTHOR_NAME_ROW = 5
AUTHOR_NAME_COL = 19

# S6
TRAINING_COUNT_ROW = 6
TRAINING_COUNT_COL = 19

# S7
WEIGHT_UNIT_ROW = 7
WEIGHT_UNIT_COL = 19

# R10 : R17
LIFTING_CATEGORY_ROW_START = 10
LIFTING_CATEGORY_ROW_END = 17
LIFTING_CATEGORY_COL = 18

# B10 : B100
ROUTINE_TABLE_TRAINING_DAY_ROW_START = 10
ROUTINE_TABLE_TRAINING_DAY_ROW_END = 100
ROUTINE_TABLE_TRAINING_DAY_ROW_GAP_DEFAULT = 18
ROUTINE_TABLE_TRAINING_DAY_COL = 2

# C10 : C112
ROUTINE_TABLE_LC_ROW_START = 10
ROUTINE_TABLE_LC_ROW_END = 112
ROUTINE_TABLE_LC_ROW_GAP_DEFAULT = 6
ROUTINE_TABLE_LC_COL = 3

# D10 : D117
ROUTINE_TABLE_REPS_ROW_START = 10
ROUTINE_TABLE_REPS_ROW_END = 117
ROUTINE_TABLE_REPS_ROW_GAP_DEFAULT = 6
ROUTINE_TABLE_REPS_COL = 4

# E10 : P116
ROUTINE_TABLE_SESSION_ROW_START = 10
ROUTINE_TABLE_SESSION_ROW_END = 117
ROUTINE_TABLE_SESSION_ROW_GAP_DEFAULT = 6
ROUTINE_TABLE_SESSION_COL_START = 5
ROUTINE_TABLE_SESSION_COL_END = 16

# R20 : U25
AE_TABLE_ROW_START = 20
AE_TABLE_ROW_END = 25
AE_TABLE_YN_COL = 19
AE_TABLE_CATEGORY_COL = 20
AE_TABLE_VOLUME_COL = 21


def bal_load_excel(dir) :
    
    workbook = load_workbook(filename=dir, read_only=True)
    
    return workbook['Schedules']


def bal_get_routine_name(sheet) :
    
    return sheet.cell(row = ROUTINE_NAME_ROW, column = ROUTINE_NAME_COL).value


def bal_get_author(sheet) :
    
    return sheet.cell(row = AUTHOR_NAME_ROW, column = AUTHOR_NAME_COL).value
    

def bal_get_training_per_weeks(sheet) :
    
    return sheet.cell(row = TRAINING_COUNT_ROW, column = TRAINING_COUNT_COL).value


def bal_get_weight_unit(sheet) :
    
    return sheet.cell(row = WEIGHT_UNIT_ROW, column = WEIGHT_UNIT_COL).value


def bal_get_lifting_category(sheet) :
    
    lifting_category = []

    for row_index in range(LIFTING_CATEGORY_ROW_START, LIFTING_CATEGORY_ROW_END+1) :
        contents = sheet.cell(row = row_index, column = LIFTING_CATEGORY_COL).value
        if contents is None :
            break
        else :
            lifting_category.append(contents)

    return lifting_category    


def bal_get_lifting_set(sheet, position) :
    
    lifting_set = []
    
    category_row = ROUTINE_TABLE_LC_ROW_START + ROUTINE_TABLE_LC_ROW_GAP_DEFAULT * position
    reps_start_row = ROUTINE_TABLE_REPS_ROW_START + ROUTINE_TABLE_REPS_ROW_GAP_DEFAULT * position
    session_start_row = ROUTINE_TABLE_SESSION_ROW_START + ROUTINE_TABLE_SESSION_ROW_GAP_DEFAULT * position

    lifting_set.append(sheet.cell(row = category_row, column = ROUTINE_TABLE_LC_COL).value)
    
    reps = []
    for index in range(reps_start_row, reps_start_row + ROUTINE_TABLE_REPS_ROW_GAP_DEFAULT) :
        contents = sheet.cell(row = index, column = ROUTINE_TABLE_REPS_COL).value
        if contents is None :
            break
        else :
            reps.append(contents)
    lifting_set.append(reps)

    session = []
    for row in range(session_start_row, session_start_row + ROUTINE_TABLE_SESSION_ROW_GAP_DEFAULT) :
        if sheet.cell(row = row, column = ROUTINE_TABLE_SESSION_COL_START).value is None :
            break
        session_list = []
        for col in range(ROUTINE_TABLE_SESSION_COL_START, ROUTINE_TABLE_SESSION_COL_END + 1) :
            contents = sheet.cell(row = row, column = col).value
            if contents is None :
                break
            else :
                session_list.append(contents)
        session.append(session_list)
        
    lifting_set.append(session)

    return lifting_set
    

def bal_get_assistance_exercise(sheet, day) :
    row = AE_TABLE_ROW_START + day
    
    if sheet.cell(row = row, column = AE_TABLE_YN_COL).value == 'N' :
        
        return None

    else :
        
        keys = ['Category', 'Volume']
        
        assistance_exercise = dict.fromkeys(keys)
        assistance_exercise['Category'] = sheet.cell(row = row, column = AE_TABLE_CATEGORY_COL).value
        assistance_exercise['Volume'] = sheet.cell(row = row, column = AE_TABLE_VOLUME_COL).value

        return assistance_exercise


def bal_get_training_days(sheet) :
   
    training_days = []
    
    for index in range(0,7) :
        row = ROUTINE_TABLE_TRAINING_DAY_ROW_START + index * ROUTINE_TABLE_TRAINING_DAY_ROW_GAP_DEFAULT
        col = ROUTINE_TABLE_TRAINING_DAY_COL
        contents = sheet.cell(row = row, column = col).value
        if contents is not None :
            training_days.append(contents)
        else :
            break
    
    return training_days
